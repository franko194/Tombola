import csv
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Assignment, Participant, SessionModel, Team, TeamMember, utc_now
from app.schemas import ParticipantCreate, ParticipantOut, ParticipantUpdate

router = APIRouter(tags=["participants"])


def ensure_session(db: Session, session_id: int) -> SessionModel:
    session = db.get(SessionModel, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    return session


def reset_generated_work(db: Session, session_id: int) -> None:
    db.query(Assignment).filter(Assignment.session_id == session_id).delete()
    db.query(TeamMember).filter(TeamMember.team_id.in_(db.query(Team.id).filter(Team.session_id == session_id))).delete()
    db.query(Team).filter(Team.session_id == session_id).delete()
    session = db.get(SessionModel, session_id)
    if session:
        session.status = "draft"
        session.updated_at = utc_now()


@router.post("/sessions/{session_id}/participants", response_model=ParticipantOut)
def create_participant(
    session_id: int,
    payload: ParticipantCreate,
    db: Session = Depends(get_db),
) -> Participant:
    ensure_session(db, session_id)
    reset_generated_work(db, session_id)
    participant = Participant(session_id=session_id, name=payload.name.strip(), ai_level=payload.ai_level)
    db.add(participant)
    db.commit()
    db.refresh(participant)
    return participant


@router.get("/sessions/{session_id}/participants", response_model=list[ParticipantOut])
def list_participants(session_id: int, db: Session = Depends(get_db)) -> list[Participant]:
    ensure_session(db, session_id)
    return list(db.query(Participant).filter(Participant.session_id == session_id).order_by(Participant.id).all())


@router.patch("/participants/{participant_id}", response_model=ParticipantOut)
def update_participant(
    participant_id: int,
    payload: ParticipantUpdate,
    db: Session = Depends(get_db),
) -> Participant:
    participant = db.get(Participant, participant_id)
    if not participant:
        raise HTTPException(status_code=404, detail="Participant not found")
    reset_generated_work(db, participant.session_id)
    data = payload.model_dump(exclude_unset=True)
    if "name" in data and data["name"] is not None:
        participant.name = data["name"].strip()
    if "ai_level" in data and data["ai_level"] is not None:
        participant.ai_level = data["ai_level"]
    participant.updated_at = utc_now()
    db.commit()
    db.refresh(participant)
    return participant


@router.delete("/participants/{participant_id}")
def delete_participant(participant_id: int, db: Session = Depends(get_db)) -> dict[str, bool]:
    participant = db.get(Participant, participant_id)
    if not participant:
        raise HTTPException(status_code=404, detail="Participant not found")
    session_id = participant.session_id
    db.delete(participant)
    reset_generated_work(db, session_id)
    db.commit()
    return {"ok": True}


def normalize_level(value: object) -> int:
    try:
        level = int(str(value).strip())
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"Invalid nivelIA value: {value}")
    if level < 0 or level > 5:
        raise HTTPException(status_code=400, detail=f"nivelIA must be between 0 and 5: {value}")
    return level


@router.post("/sessions/{session_id}/participants/import", response_model=list[ParticipantOut])
async def import_participants(
    session_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> list[Participant]:
    ensure_session(db, session_id)
    content = await file.read()
    rows: list[dict[str, object]] = []

    filename = (file.filename or "").lower()
    if filename.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(max_row=1))]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, values)))
    else:
        text = content.decode("utf-8-sig")
        rows = list(csv.DictReader(StringIO(text)))

    created: list[Participant] = []
    reset_generated_work(db, session_id)
    for row in rows:
        name = str(row.get("nombre") or row.get("name") or "").strip()
        raw_level = row.get("nivelIA", row.get("ai_level"))
        if not name:
            raise HTTPException(status_code=400, detail="Imported rows must include nombre")
        participant = Participant(session_id=session_id, name=name, ai_level=normalize_level(raw_level))
        db.add(participant)
        created.append(participant)

    db.commit()
    for participant in created:
        db.refresh(participant)
    return created
